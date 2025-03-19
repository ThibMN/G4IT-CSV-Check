import openpyxl

class XlsxHandler:
    """Simple handler for XLSX file operations.

    Provides methods to read, write, and append data to XLSX files 
    with automatic header management.
    """


    def __init__(self, file):
        """Initializes the Xlsx class with a file path.

        Args:
            file (String): Path to the XLSX file to be manipulated.
        """
        self.file = file


    def load_data(self):
        """Reads data from XLSX file.

        Returns:
            list: List of dictionaries where each dictionary represents a row.
        """
        try:
            wb = openpyxl.load_workbook(self.file)
            sheet = wb.active
            data = []
            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(headers, row)))

            return data
        except FileNotFoundError:
            print(f"Erreur: Le fichier '{self.file}' est introuvable.")
            return []
        except PermissionError:
            print(f"Erreur: Pas d'autorisation pour accéder au fichier '{self.file}'.")
            return []
        except openpyxl.utils.exceptions.InvalidFileException:
            print(f"Erreur: Le fichier '{self.file}' n'est pas un fichier Excel valide.")
            return []
        except Exception as e:
            print(f"Erreur inattendue lors du chargement du fichier '{self.file}': {str(e)}")
            return []


    def write_data(self, data, header=None):
        """Writes data to XLSX file.

        Args:
            data (list): List of dictionaries to write.
            header (list, optional): Column headers. Defaults to None.
        """
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active

            if header is None:
                header = list(data[0].keys()) if data else []

            sheet.append(header)

            for row in data:
                sheet.append([row.get(col) for col in header])

            wb.save(self.file)
        except PermissionError:
            print(f"Erreur: Pas d'autorisation pour modifier le fichier '{self.file}'.")
        except IOError as e:
            print(f"Erreur d'entrée/sortie lors de l'écriture dans '{self.file}': {str(e)}")
        except Exception as e:
            print(f"Erreur inattendue lors de la modification des données dans '{self.file}': {str(e)}")

    def append_row(self, line):
        """Appends a single line (row) to XLSX file.

        Args:
            line (dict): Dictionary representing a row to add.
        """
        try:
            wb = openpyxl.load_workbook(self.file)
            sheet = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(line.keys())  # Write header if file does not exist
        except PermissionError:
            print(f"Erreur: Pas d'autorisation pour accéder au fichier '{self.file}'.")
            return
        except openpyxl.utils.exceptions.InvalidFileException:
            print(f"Erreur: Le fichier '{self.file}' n'est pas un fichier Excel valide.")
            return
        except Exception as e:
            print(f"Erreur inattendue lors de l'ouverture du fichier '{self.file}': {str(e)}")
            return

        try:
            sheet.append(line.values())
            wb.save(self.file)
        except PermissionError:
            print(f"Erreur: Pas d'autorisation pour modifier le fichier '{self.file}'.")
        except IOError as e:
            print(f"Erreur d'entrée/sortie lors de l'ajout de ligne dans '{self.file}': {str(e)}")
        except Exception as e:
            print(f"Erreur inattendue lors de l'ajout de ligne dans '{self.file}': {str(e)}")

    def validate_dates(self, date_column):
        """
        Checks if dates in XLSX are in correct mm-dd-yyyy format.
        
        Args:
            date_column (str): Name of the column containing dates

        Returns:
            bool: True if dates need switching (month > 12 found), False otherwise
        """
        data = self.load_data()
        invalid_rows = []
        
        for i, row in enumerate(data, 1):
            if date_column not in row:
                raise KeyError(f"La colonne '{date_column}' est absente du fichier XLSX. Vérifiez l'orthographe ou les en-têtes du fichier.")

            try:
                if not row[date_column]:  # Skip empty cells
                    invalid_rows.append(f"Ligne {i}: valeur de date manquante")
                    continue

                date_str = str(row[date_column])

                if '/' in date_str:
                    month, day, year = map(int, date_str.split('/'))
                else:
                    month, day, year = map(int, date_str.split('-'))

                if month > 12:
                    print(f"Format de date inversé détecté à la ligne {i}: {date_str} (mois={month} > 12)")
                    return True

            except ValueError:
                invalid_rows.append(f"Ligne {i}: format de date invalide '{row[date_column]}' (doit être JJ-MM-AAAA ou MM-JJ-AAAA)")
                continue
            except AttributeError:
                invalid_rows.append(f"Ligne {i}: type de donnée invalide pour la date '{type(row[date_column])}'")
                continue

        if invalid_rows:
            print(f"Attention: {len(invalid_rows)} lignes avec des dates non valides trouvées:")
            for msg in invalid_rows[:5]:
                print(f"  - {msg}")
            if len(invalid_rows) > 5:
                print(f"  ... et {len(invalid_rows)-5} autres problèmes.")

        return False

    def fix_dates(self, date_column):
        """
        Fixes dates by switching month and day if month > 12.
        Also standardizes separator to hyphen.

        Args:
            date_column (str): Name of the column containing dates

        Returns:
            list: Corrected data with fixed dates
        """
        data = self.load_data()
        unfixable_rows = []
        fixed_count = 0
        
        for i, row in enumerate(data, 1):
            if date_column not in row:
                raise KeyError(f"La colonne '{date_column}' est absente du fichier XLSX. Impossible de corriger les dates.")

            try:
                if not row[date_column]:  # Skip empty cells
                    continue

                date_str = str(row[date_column])

                if '/' in date_str:
                    month, day, year = map(int, date_str.split('/'))
                    separator_fixed = True
                else:
                    month, day, year = map(int, date_str.split('-'))
                    separator_fixed = False

                if month > 12:
                    # Swap month and day and standardize to hyphen
                    row[date_column] = f"{day:02d}-{month:02d}-{year}"
                    fixed_count += 1
                    print(f"Corrigé: '{date_str}' → '{row[date_column]}' (ligne {i})")
                elif separator_fixed:
                    # Standardize to hyphen without swapping
                    row[date_column] = f"{month:02d}-{day:02d}-{year}"
                    print(f"Normalisé: '{date_str}' → '{row[date_column]}' (ligne {i})")

            except ValueError:
                unfixable_rows.append(f"Ligne {i}: impossible de parser '{row[date_column]}'")
                continue
            except AttributeError:
                unfixable_rows.append(f"Ligne {i}: type de donnée invalide pour la date '{type(row[date_column])}'")
                continue
        
        if unfixable_rows:
            print(f"\nAttention: {len(unfixable_rows)} lignes n'ont pas pu être corrigées:")
            for msg in unfixable_rows[:5]:
                print(f"  - {msg}")
            if len(unfixable_rows) > 5:
                print(f"  ... et {len(unfixable_rows)-5} autres problèmes.")
        
        print(f"\nRésumé: {fixed_count} dates inversées corrigées, {len(unfixable_rows)} lignes non traitables.")
        return data