import openpyxl

class XlsxHandler:
    """Simple handler for XLSX file operations.

    Provides methods to read, write, and append data to XLSX files 
    with automatic header management.
    """


    def __init__(self, file):
        """Initializes the Xlsx class with a file path.

        Args:
            file (String): Path to the XLSX file to be manipulated.
        """
        self.file = file


    def load_data(self):
        """Reads data from XLSX file.

        Returns:
            list: List of dictionaries where each dictionary represents a row.
        """
        wb = openpyxl.load_workbook(self.file)
        sheet = wb.active
        data = []
        headers = [cell.value for cell in sheet[1]]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))

        return data


    def write_data(self, data, header=None):
        """Writes data to XLSX file.

        Args:
            data (list): List of dictionaries to write.
            header (list, optional): Column headers. Defaults to None.
        """

        wb = openpyxl.Workbook()
        sheet = wb.active

        if header is None:
            header = list(data[0].keys()) if data else []

        sheet.append(header)

        for row in data:
            sheet.append([row.get(col) for col in header])

        wb.save(self.file)

    def append_row(self, line):
        """Appends a single line (row) to XLSX file.

        Args:
            line (dict): Dictionary representing a row to add.
        """
        try:
            wb = openpyxl.load_workbook(self.file)
            sheet = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(line.keys())  # Write header if file does not exist

        sheet.append(line.values())
        wb.save(self.file)

    def validate_dates(self, date_column):
        """
        Checks if dates in XLSX are in correct mm-dd-yyyy format.
        
        Args:
            date_column (str): Name of the column containing dates

        Returns:
            bool: True if dates need switching (month > 12 found), False otherwise
        """
        data = self.load_data()
        
        for row in data:
            if date_column not in row:
                raise KeyError(f"Column '{date_column}' not found in XLSX")

            try:
                if not row[date_column]:  # Skip empty cells
                    continue

                date_str = str(row[date_column])

                if '/' in date_str:
                    month, day, year = map(int, date_str.split('/'))

                else:
                    month, day, year = map(int, date_str.split('-'))

                if month > 12:
                    return True

            except (ValueError, AttributeError):
                continue

        return False

    def fix_dates(self, date_column):
        """
        Fixes dates by switching month and day if month > 12.
        Also standardizes separator to hyphen.

        Args:
            date_column (str): Name of the column containing dates

        Returns:
            list: Corrected data with fixed dates
        """
        data = self.load_data()
        
        for row in data:
            try:
                if not row[date_column]:  # Skip empty cells
                    continue

                date_str = str(row[date_column])

                if '/' in date_str:
                    month, day, year = map(int, date_str.split('/'))

                else:
                    month, day, year = map(int, date_str.split('-'))

                if month > 12:
                    # Swap month and day and standardize to hyphen
                    row[date_column] = f"{year}-{day:02d}-{month:02d}"
                else:
                    # Standardize to hyphen without swapping
                    row[date_column] = f"{year}-{month:02d}-{day:02d}"
            except (ValueError, AttributeError):
                continue

        return data